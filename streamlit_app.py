"""
Solution Unifiée - Forms 2050, 2051, 2052, 2053
Détection automatique + Extraction + Génération Excel multi-années
"""

import io
import re
from typing import Optional, Tuple, List, Dict
import pdfplumber
import streamlit as st
from PIL import Image, ImageDraw, ImageFont
from openpyxl import load_workbook

if "manual_scopes" not in st.session_state:
    # dict clé = (pdf_name, form) -> (X1, X2)
    st.session_state["manual_scopes"] = {}


# ============================ CONFIG ============================
st.set_page_config(page_title="Extracteur Forms 2050-2053", layout="wide")
st.title("🎯 Agent Extracteur de liasses fiscales")

# ============================ EXCEL GENERATION ============================
# Colonnes cibles (Excel est 1-indexé) : G=7, H=8, I=9
COL_G, COL_H, COL_I = 7, 8, 9
HEADER_ROW = 2  # ligne de l'en-tête de dates (ex: "31/12/2024")

def assign_years_to_fixed_columns(years_sorted_asc):
    """
    Mappe les années vers des colonnes fixes (jusqu'à 3 années) :
      - la plus ancienne -> I
      - la suivante     -> H
      - la plus récente -> G
    years_sorted_asc : liste d'années triées croissantes (ex: [2022, 2023, 2024])
    """
    cols_by_order = [COL_I, COL_H, COL_G]  # ordre ancien -> récent
    return {y: cols_by_order[i] for i, y in enumerate(years_sorted_asc[:3])}

def write_header_dates(ws, year_to_col):
    """
    Écrit les dates d'en-tête "31/12/AAAA" sur la ligne HEADER_ROW
    aux colonnes cibles calculées.
    """
    for year, col in year_to_col.items():
        ws.cell(row=HEADER_ROW, column=col).value = f"31/12/{year}"

def build_code_row_map(ws, code_col_candidates=range(1, 15)):
    """
    Construit un mapping {code: ligne} en scannant les premières colonnes (par défaut 1..14).
    
    IMPORTANT: Cette version cherche TOUS les codes dans ALL_CODES,
    sans filtrage par feuille. C'est ce qui permet de trouver les codes
    Form 2051 même s'ils sont mélangés avec d'autres codes.
    """
    row_map, max_row = {}, ws.max_row
    for r in range(1, max_row + 1):
        for c in code_col_candidates:
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str):
                t = v.strip()
                # Chercher dans ALL_CODES, pas de filtre spécifique
                if t in ALL_CODES and t not in row_map:
                    row_map[t] = r
    return row_map

def write_mapping_into_sheet(ws, mapping, code_to_row, target_col, filter_codes):
    """
    Écrit les valeurs d'un 'mapping' {code: valeur} dans la feuille 'ws' :
      - target_col : colonne année (G/H/I selon l'année)
      - filter_codes : limite aux codes pertinents pour cette feuille
    """
    for code in filter_codes:
        row = code_to_row.get(code)
        if not row:
            continue
        val = mapping.get(code)
        if val is None:
            ws.cell(row=row, column=target_col).value = None
        else:
            # Convertir en nombre (gérer les négatifs)
            try:
                val_clean = str(val).replace(",", ".").replace(" ", "")
                num_val = float(val_clean)
                ws.cell(row=row, column=target_col).value = num_val
            except:
                ws.cell(row=row, column=target_col).value = val

def fill_excel(model_bytes, year_to_mapping,
               bilan_sheet="Saisie Bilan",
               cr_sheet="Saisie Cpte Res.",
               codes_bilan=None,
               codes_cr=None):
    """
    Remplit le modèle Excel à partir des extractions.

    Entrées :
    - model_bytes : contenu binaire du fichier modèle (.xlsx)
    - year_to_mapping : dict {année:int -> {code:str -> valeur:str|None}}
        ex: {2024: {"AA": "12345", "AB": None, ...}, 2023: {...}, 2022: {...}}
    - bilan_sheet : nom de l'onglet bilan (codes 'liasse')
    - cr_sheet    : nom de l'onglet compte de résultat (codes 'saisie')
    - codes_bilan : set des codes attendus pour 'Saisie Bilan'
    - codes_cr    : set des codes attendus pour 'Saisie Cpte Res.'

    Sortie :
    - bytes du fichier Excel complété
    """
    # Ouverture du modèle
    bio = io.BytesIO(model_bytes)
    wb = load_workbook(bio)

    if bilan_sheet not in wb.sheetnames or cr_sheet not in wb.sheetnames:
        raise ValueError("Feuilles manquantes (Saisie Bilan / Saisie Cpte Res.).")

    ws_b, ws_cr = wb[bilan_sheet], wb[cr_sheet]

    # Colonnes années : plus ancienne -> I, suivante -> H, récente -> G
    years_sorted = sorted(year_to_mapping.keys())
    if not years_sorted:
        # Rien à écrire, renvoyer le modèle tel quel
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    year_to_col = assign_years_to_fixed_columns(years_sorted)

    # En-têtes de dates
    write_header_dates(ws_b, year_to_col)
    write_header_dates(ws_cr, year_to_col)

    # Indexation des lignes par code
    rows_b = build_code_row_map(ws_b)
    rows_c = build_code_row_map(ws_cr)

    # Écriture des valeurs pour chaque année
    for year, mapping in year_to_mapping.items():
        col = year_to_col.get(year)
        if not col:
            continue
        
        if codes_bilan:
            write_mapping_into_sheet(ws_b, mapping, rows_b, col, codes_bilan)
        else:
            write_mapping_into_sheet(ws_b, mapping, rows_b, col, set(mapping.keys()))
        
        if codes_cr:
            write_mapping_into_sheet(ws_cr, mapping, rows_c, col, codes_cr)
        else:
            write_mapping_into_sheet(ws_cr, mapping, rows_c, col, set(mapping.keys()))

    # Sauvegarde
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

# ============================ CODES PAR FORM ============================
CODES_2050 = [
    "AA", "AB", "AD", "AF", "AH", "AJ", "AL", "AN", "AP", "AR", "AT", "AV", "AX", 
    "CS", "CU", "BB", "BD", "BF", "BH", "BJ", "BL", "BN", "BP", "BR", "BT", "BV", 
    "BX", "BZ", "CB", "CD", "CF", "CH", "CJ", "CL", "CM", "CN", "CO"
]

CODES_2051 = [
    "DA", "DB", "DC", "DD", "DE", "DF", "DG", "DH", "DI", "DJ", "DK", "DL",
    "DM", "DN", "DO", "DP", "DQ", "DR", "DS", "DT", "DU", "DV", "DW", "DX",
    "DY", "DZ", "EA", "EB", "EC", "ED", "EE", "1B", "1C", "1D", "1E", "EF",
    "EG", "EH"
]

CODES_2052 = [
    "FC", "FF", "FI", "FL", "FM", "FN", "FO", "FP", "FQ", "FR", "FS", "FT",
    "FU", "FV", "FW", "FX", "FY", "FZ", "GA", "GB", "GC", "GD", "GE", "GF",
    "GG", "GH", "GI", "GJ", "GK", "GL", "GM", "GN", "GO", "GP", "GQ", "GR",
    "GS", "GT", "GU", "GV", "GW"
]

CODES_2053 = [
    "HA", "HB", "HC", "HD", "HE", "HF", "HG", "HH", "HI", "HJ", "HK", "HL",
    "HM", "HN", "YY", "YZ"
]

# ✅ CORRECTION DU ROUTAGE ICI
CODES_BILAN_ALL = set(CODES_2050 + CODES_2051)   # -> va sur "Saisie Bilan"
CODES_CR_ALL    = set(CODES_2052 + CODES_2053)   # -> va sur "Saisie Cpte Res."

# Tous les codes pour build_code_row_map
ALL_CODES = CODES_BILAN_ALL | CODES_CR_ALL

CODES_BY_FORM = {
    "2050": CODES_2050,
    "2051": CODES_2051,
    "2052": CODES_2052,
    "2053": CODES_2053
}

# ============================ DÉTECTION FORMULAIRE ============================
def find_form_from_text(text: str) -> Optional[str]:
    if re.search(r'DGFiP\s*N[°º]\s*2050', text): return "2050"
    if re.search(r'DGFiP\s*N[°º]\s*2051', text): return "2051"
    if re.search(r'DGFiP\s*N[°º]\s*2052', text): return "2052"
    if re.search(r'DGFiP\s*N[°º]\s*2053', text): return "2053"
    return None

# ============================ LIGNES VERTICALES ============================
def extract_vertical_lines(page) -> List[dict]:
    verts = []
    
    for ln in page.lines:
        if abs(ln["x0"] - ln["x1"]) <= 1.0:
            verts.append({
                "x": (ln["x0"] + ln["x1"]) / 2,
                "y0": min(ln["y0"], ln["y1"]),
                "y1": max(ln["y0"], ln["y1"]),
                "height": abs(ln["y1"] - ln["y0"])
            })
    
    for cv in page.curves:
        if abs(cv["x0"] - cv["x1"]) <= 1.0:
            verts.append({
                "x": (cv["x0"] + cv["x1"]) / 2,
                "y0": min(cv["y0"], cv["y1"]),
                "y1": max(cv["y0"], cv["y1"]),
                "height": abs(cv["y1"] - cv["y0"])
            })
    
    for rc in page.rects:
        x0, y0, x1, y1 = rc["x0"], rc["y0"], rc["x1"], rc["y1"]
        height = abs(y1 - y0)
        verts.append({"x": x0, "y0": min(y0, y1), "y1": max(y0, y1), "height": height})
        verts.append({"x": x1, "y0": min(y0, y1), "y1": max(y0, y1), "height": height})
    
    return verts

def merge_close_vertical_lines(lines: List[dict], tolerance: float = 2.0) -> List[dict]:
    if not lines:
        return []
    
    sorted_lines = sorted(lines, key=lambda l: l["x"])
    merged = [sorted_lines[0]]
    
    for line in sorted_lines[1:]:
        last = merged[-1]
        if abs(line["x"] - last["x"]) <= tolerance:
            merged[-1] = {
                "x": (last["x"] + line["x"]) / 2,
                "y0": min(last["y0"], line["y0"]),
                "y1": max(last["y1"], line["y1"]),
                "height": max(last["y1"], line["y1"]) - min(last["y0"], line["y0"])
            }
        else:
            merged.append(line)
    
    return merged

# ============================ FORM 2050 ============================
def detect_exercice_n_header_2050(page) -> Optional[Tuple[float, float]]:
    """
    Objectif : localiser horizontalement la colonne "Exercice N / Net 3"
    et PAS "N-1 / Net 4".

    Nouvelle logique :
    - On lit l'en-tête de gauche à droite.
    - On collecte tous les mots qui ressemblent à "Exercice N", "Net", "Net 3",
      "Exercice N clos le".
    - On EXCLUT tout ce qui mentionne "N-1", "N - 1", "N-2", etc. ou "4".
    - On prend le candidat VALIDE le plus à gauche.
    - On renvoie (x0-30, x1+30) de ce candidat comme fenêtre approximative.
    """

    words = page.extract_words(use_text_flow=False, x_tolerance=3, y_tolerance=3)

    # On limite la recherche à la bande supérieure de la page (zone d'entête)
    header_zone_y = page.height * 0.40

    # regex utilitaires
    re_excl = re.compile(
        r'(n\s*-\s*1|n\s*-\s*2|n-1|n-2|exercice\s*n\s*-\s*1|net\s*4|net\s*\(?4\)?|exercice\s*n-1|n\s*-?\s*1)',
        re.IGNORECASE
    )

    # inclusions "positives" -> Exercice N (courant)
    re_incl_list = [
        re.compile(r'exercice\s*n(?!\s*-\s*\d)', re.IGNORECASE),     # "Exercice N" mais pas "N-1"
        re.compile(r'net\s*3', re.IGNORECASE),                        # "Net 3"
        re.compile(r'\bnet\b$', re.IGNORECASE),                       # juste "Net"
    ]

    # aussi, les cellules de date "30 06 2018" etc. sont dans la même colonne N.
    # On peut exploiter les dates "JJ MM AAAA" ou "JJ/MM/AAAA"
    date_pattern = re.compile(r'(?:[0-3]?\d)[/\s]?(?:0?\d)[/\s]?\d{4}')

    # On construit la liste des candidats valides
    candidates = []

    for w in words:
        # ignorer si trop bas
        if w["top"] > header_zone_y:
            continue

        txt = (w.get("text") or "").strip()

        if not txt:
            continue

        lower_txt = txt.lower()

        # hard exclude si on voit N-1 etc.
        if re_excl.search(lower_txt):
            continue

        # match positif ?
        positive = any(r.search(lower_txt) for r in re_incl_list)

        # ou bien une date ? (souvent juste sous "Exercice N clos le")
        # mais attention : une date seule peut aussi être sous N-1,
        # donc on ne prend la date QUE si on n'a pas trouvé mieux après.
        is_date_like = bool(date_pattern.search(lower_txt))

        if positive or is_date_like:
            candidates.append({
                "text": txt,
                "x0": w["x0"],
                "x1": w["x1"],
                "y": w["top"],
                "is_date_like": is_date_like,
                "positive": positive,
            })

    if not candidates:
        return None

    # Priorité aux candidats "positifs" (Exercice N / Net / Net 3),
    # parce que les dates seules peuvent apparaître sous plusieurs colonnes.
    positives = [c for c in candidates if c["positive"]]
    if positives:
        # TRÈS IMPORTANT : on prend le plus à GAUCHE (x0 minimal), pas le plus à droite
        best = min(positives, key=lambda c: c["x0"])
    else:
        # fallback : on n'a que des dates propres sans "N-1", on prend la plus à gauche
        best = min(candidates, key=lambda c: c["x0"])

    # on élargit un peu la boîte pour donner une marge à find_column_scope_2050
    return (best["x0"] - 30, best["x1"] + 30)

def find_column_scope_2050(page) -> Optional[Tuple[float, float]]:
    """
    On récupère la position horizontale estimée de la colonne "Exercice N / Net 3"
    via detect_exercice_n_header_2050.

    Puis on cherche les 2 lignes verticales longues les plus proches
    autour de ce centre pour délimiter la colonne.
    """

    header_pos = detect_exercice_n_header_2050(page)
    if not header_pos:
        return None

    header_x_min, header_x_max = header_pos
    header_center = (header_x_min + header_x_max) / 2

    verts = extract_vertical_lines(page)
    if not verts:
        return None

    verts = merge_close_vertical_lines(verts, tolerance=2.0)

    # On ne garde que les lignes verticales longues (celles qui dessinent les colonnes)
    min_height = page.height * 0.30
    long_verts = [v for v in verts if v["height"] >= min_height]

    if len(long_verts) < 2:
        return None

    # ligne juste à gauche du header_center
    left_candidates = [v for v in long_verts if v["x"] < header_center]
    if not left_candidates:
        return None
    X1 = max(left_candidates, key=lambda v: v["x"])["x"]

    # ligne juste à droite du header_center
    right_candidates = [v for v in long_verts if v["x"] > header_center]
    if not right_candidates:
        return None
    X2 = min(right_candidates, key=lambda v: v["x"])["x"]

    # sanity check largeur colonne
    width = X2 - X1
    if width < 25 or width > page.width * 0.5:
        return None

    return (X1, X2)

# ============================ AIDE: COLONNE À DROITE DES CODES ============================
def find_column_scope_right_of_codes(page, code_regex: str, codes_list: List[str]) -> Optional[Tuple[float, float]]:
    """
    Détermine la colonne immédiatement à DROITE de la colonne des codes.
    - On calcule la moyenne des x1 des codes (bord droit des codes).
    - On prend la première longue ligne verticale à droite comme X1,
      puis la suivante comme X2.
    - Fallback: fenêtre [avg_code_x + 10, + 200].
    """
    words = page.extract_words(use_text_flow=False, x_tolerance=2, y_tolerance=2)
    rx = re.compile(code_regex)

    code_rights = []
    for w in words:
        t = (w.get("text") or "").strip()
        if t in codes_list and rx.match(t):
            code_rights.append(w["x1"])

    if not code_rights:
        return None

    avg_code_x = sum(code_rights) / len(code_rights)

    verts = extract_vertical_lines(page)
    if verts:
        verts = merge_close_vertical_lines(verts, tolerance=2.0)
        min_height = page.height * 0.30
        long_verts = [v for v in verts if v["height"] >= min_height]
        long_verts = sorted(long_verts, key=lambda v: v["x"])

        # Chercher la première ligne à droite des codes
        right_of_codes = [v for v in long_verts if v["x"] > avg_code_x + 2]
        if len(right_of_codes) >= 2:
            X1 = right_of_codes[0]["x"]
            X2 = right_of_codes[1]["x"]
            # Sanity checks
            if 25 <= (X2 - X1) <= page.width * 0.5:
                return (X1, X2)

    # Fallback si les lignes verticales ne sont pas exploitables
    return (avg_code_x + 10, min(page.width - 5, avg_code_x + 200))

# ============================ FORM 2051 ============================
# ⚠️ NOUVELLE RÈGLE: on prend la colonne immédiatement à droite des codes (Exercice N)
def find_column_scope_2051(page) -> Optional[Tuple[float, float]]:
    # Codes 2051: deux lettres ou 1B..1E
    return find_column_scope_right_of_codes(
        page,
        code_regex=r'^([A-Z]{2}|1[B-E])$',
        codes_list=CODES_2051
    )

# ============================ FORM 2052 ============================
def detect_total_header_2052(page) -> Optional[Tuple[float, float]]:
    words = page.extract_words(use_text_flow=False, x_tolerance=3, y_tolerance=3)
    header_zone_y = page.height * 0.40
    
    # BANNIR : France, Export, N-1, etc.
    exclude_patterns = [
        r'france', r'export', r'livraison', r'intracommunautaire',
        r'n\s*-\s*1', r'n-1', r'précédent', r'antérieur'
    ]
    candidates = []
    
    for w in words:
        if w["top"] > header_zone_y:
            continue
        text = w.get("text", "").lower().strip()
        is_excluded = any(re.search(pat, text) for pat in exclude_patterns)
        if is_excluded:
            continue
        if text == "total":
            candidates.append({
                "text": w["text"],
                "x0": w["x0"],
                "x1": w["x1"],
                "y": w["top"]
            })
    
    if not candidates:
        return None
    
    # TOUJOURS prendre le candidat le plus à droite
    best = max(candidates, key=lambda c: c["x0"])
    return (best["x0"], best["x1"])

def find_column_scope_2052(page) -> Optional[Tuple[float, float]]:
    header_pos = detect_total_header_2052(page)
    if not header_pos:
        return None
    
    header_x_min, header_x_max = header_pos
    header_center = (header_x_min + header_x_max) / 2
    
    verts = extract_vertical_lines(page)
    if not verts:
        return None
    
    verts = merge_close_vertical_lines(verts, tolerance=2.0)
    min_height = page.height * 0.30
    long_verts = [v for v in verts if v["height"] >= min_height]
    
    if len(long_verts) < 2:
        return None
    
    left_candidates = [v for v in long_verts if v["x"] < header_center]
    if not left_candidates:
        return None
    X1 = max(left_candidates, key=lambda v: v["x"])["x"]
    
    right_candidates = [v for v in long_verts if v["x"] > header_center]
    if not right_candidates:
        return None
    X2 = min(right_candidates, key=lambda v: v["x"])["x"]
    
    if X2 - X1 < 30 or X2 - X1 > page.width * 0.5:
        return None
    
    return (X1, X2)

# ============================ FORM 2053 ============================
# ⚠️ NOUVELLE RÈGLE: idem 2051 — colonne immédiatement à droite des codes (Exercice N)
def find_column_scope_2053(page) -> Optional[Tuple[float, float]]:
    # Codes 2053: deux lettres + YY/YZ
    return find_column_scope_right_of_codes(
        page,
        code_regex=r'^([A-Z]{2}|YY|YZ)$',
        codes_list=CODES_2053
    )

# ============================ DISPATCHER ============================
def find_column_scope(page, form: str) -> Optional[Tuple[float, float]]:
    if form == "2050":
        return find_column_scope_2050(page)
    elif form == "2051":
        return find_column_scope_2051(page)
    elif form == "2052":
        return find_column_scope_2052(page)
    elif form == "2053":
        return find_column_scope_2053(page)
    return None

# ============================ CODES + K ============================
def find_code_positions(page, form: str) -> Dict[str, dict]:
    words = page.extract_words(use_text_flow=False, x_tolerance=1, y_tolerance=1)
    code_positions = {}
    
    codes_list = CODES_BY_FORM.get(form, [])
    
    if form == "2051":
        code_re = re.compile(r'^([A-Z]{2}|1[B-E])$')
    elif form == "2053":
        code_re = re.compile(r'^([A-Z]{2}|YY|YZ)$')
    else:
        code_re = re.compile(r'^[A-Z]{2}$')
    
    for w in words:
        text = w.get("text", "").strip()
        if code_re.match(text) and text in codes_list:
            if text not in code_positions:
                code_positions[text] = {
                    "y": w["top"],
                    "x0": w["x0"],
                    "x1": w["x1"],
                    "bottom": w["bottom"]
                }
    
    return code_positions

def calculate_k_constant(code_positions: Dict[str, dict], form: str) -> Optional[float]:
    if len(code_positions) < 2:
        return None
    
    sorted_codes = sorted(code_positions.items(), key=lambda x: x[1]["y"])
    
    spacings = []
    for i in range(len(sorted_codes) - 1):
        y_current = sorted_codes[i][1]["y"]
        y_next = sorted_codes[i+1][1]["y"]
        spacing = y_next - y_current
        
        # 2053 plus permissif
        max_spacing = 100 if form == "2053" else 50
        
        if 10 < spacing < max_spacing:
            spacings.append(spacing)
    
    if not spacings:
        return None
    
    spacings.sort()
    median_spacing = spacings[len(spacings) // 2]
    k = median_spacing / 2.0
    
    return k

def calculate_zones_with_k(code_positions: Dict[str, dict], k: float, page) -> Dict[str, Tuple[float, float]]:
    zones = {}
    
    for code, pos in code_positions.items():
        y_code = pos["y"]
        Y1 = max(0, y_code - k + 1)
        Y2 = min(page.height, y_code + k + 1)
        zones[code] = (Y1, Y2)
    
    return zones

# ============================ EXTRACTION MONTANTS ============================
def clean_amount(text: str) -> Optional[str]:
    """
    Nettoie et formate un montant, gère les négatifs
    
    Formats négatifs:
    - (123) → -123
    - -123 → -123
    """
    # Détecter négatif
    is_negative = False
    
    # Format (123)
    if re.match(r'^\(.*\)$', text):
        is_negative = True
        text = re.sub(r'[()]', '', text)
    
    # Format -123
    if text.startswith('-'):
        is_negative = True
        text = text[1:]
    
    # Nettoyer: garder chiffres, virgules, points
    cleaned = re.sub(r'[^\d,.]', '', text)
    cleaned = cleaned.replace(" ", "")
    
    if not cleaned:
        return None
    
    # Ajouter signe négatif
    if is_negative:
        cleaned = "-" + cleaned
    
    return cleaned

def extract_amount_in_zone(page, code: str, X1: float, X2: float, Y1: float, Y2: float) -> Optional[str]:
    words = page.extract_words(use_text_flow=False, x_tolerance=1, y_tolerance=1)

    candidates = []

    for w in words:
        word_text = w.get("text", "").strip()
        word_x0 = w["x0"]
        word_x1 = w["x1"]
        word_y = w["top"]

        # chevauchement en X + rangée correcte en Y
        x_overlap = not (word_x1 < X1 or word_x0 > X2)
        y_overlap = Y1 <= word_y <= Y2

        if x_overlap and y_overlap:
            # On ne retient que les bouts qui semblent "numériques" ou formats de négatif
            if re.search(r'\d', word_text) or re.search(r'[\(\)\-]', word_text):
                candidates.append({
                    "text": word_text,
                    "x0": word_x0,
                    "x1": word_x1,
                    "y": word_y
                })

    if not candidates:
        return None

    # Trier les fragments détectés par position X (gauche -> droite)
    candidates = sorted(candidates, key=lambda c: c["x0"])

    # Regrouper les fragments contigus en "lignes"
    groups = []
    current_group = [candidates[0]]

    for i in range(1, len(candidates)):
        prev = current_group[-1]
        curr = candidates[i]

        same_line = abs(curr["y"] - prev["y"]) <= 3
        close_x = (curr["x0"] - prev["x1"]) <= 15

        if same_line and close_x:
            current_group.append(curr)
        else:
            groups.append(current_group)
            current_group = [curr]

    groups.append(current_group)

    amounts = []

    for group in groups:
        # texte complet du groupe
        full_text = " ".join([c["text"] for c in group])
        cleaned = clean_amount(full_text)
        if not cleaned:
            continue

        # position horizontale du groupe (début)
        group_x0 = min(c["x0"] for c in group)

        # version sans espaces
        cleaned_no_space = cleaned.replace(" ", "")

        # --- ANTI BRUIT SPECIAL 2051 (ex: "1" venant de "1B") ---
        # Si:
        #   - la "valeur" est minuscule (1 ou 12 etc.)
        #   - ET cette valeur est collée tout au bord gauche du scope (donc probablement un bout du code)
        # Alors on l'ignore.
        #
        # Distance horizontale du groupe par rapport au bord gauche du scope:
        dist_from_left = group_x0 - X1

        if len(cleaned_no_space) < 3 and dist_from_left < 10:
            # trop court ET trop collé à gauche => bruit, on skip
            continue

        # Sinon on accepte, même si c'est "1"
        amounts.append({
            "value": cleaned,
            "score": len(cleaned_no_space.replace("-", "")),  # sert à départager après
        })

    if not amounts:
        return None

    # On choisit le montant "le plus crédible"
    # score = longueur numérique (donc "181210249" > "1")
    # mais si on a qu'un "1" légitime (cas EH), il passera car ce sera le seul accepté
    best = max(amounts, key=lambda a: a["score"])
    return best["value"]



def extract_all_amounts(page, code_positions: Dict[str, dict], 
                       zones: Dict[str, Tuple[float, float]], 
                       column_scope: Tuple[float, float]) -> Dict[str, Optional[str]]:
    X1, X2 = column_scope
    amounts = {}
    
    for code in code_positions.keys():
        if code in zones:
            Y1, Y2 = zones[code]
            amount = extract_amount_in_zone(page, code, X1, X2, Y1, Y2)
            amounts[code] = amount
    
    return amounts

# ============================ VISUALISATION ============================
def get_saved_scope(pdf_name: str, form: str):
    """Retourne le scope manuel stocké pour (pdf_name, form) ou None."""
    return st.session_state["manual_scopes"].get((pdf_name, form))

def save_scope(pdf_name: str, form: str, X1: float, X2: float):
    """Sauvegarde/écrase le scope manuel pour (pdf_name, form)."""
    st.session_state["manual_scopes"][(pdf_name, form)] = (float(X1), float(X2))

def get_effective_scope(page, pdf_name: str, form: str):
    """
    Pour l'extraction (debug ou batch) :
    1. Si l'utilisateur a déjà sauvé un scope manuel pour (pdf_name, form), on le renvoie.
    2. Sinon on calcule le scope auto.
    """
    saved = get_saved_scope(pdf_name, form)
    if saved:
        return saved  # (X1, X2)

    auto = find_column_scope(page, form)
    return auto  # peut être None

def draw_scope_visualization(page, column_scope: Optional[Tuple[float, float]], 
                            code_positions: Dict[str, dict],
                            zones: Dict[str, Tuple[float, float]],
                            k: Optional[float],
                            highlight_code: Optional[str] = None) -> Image.Image:
    img = page.to_image(resolution=150).original
    draw = ImageDraw.Draw(img)
    
    scale_x = img.width / page.width
    scale_y = img.height / page.height
    
    if not column_scope:
        return img
    
    X1, X2 = column_scope
    x1_scaled = X1 * scale_x
    x2_scaled = X2 * scale_x
    
    # Colonne verte
    draw.line([(x1_scaled, 0), (x1_scaled, img.height)], fill=(0, 255, 0), width=4)
    draw.line([(x2_scaled, 0), (x2_scaled, img.height)], fill=(0, 255, 0), width=4)
    
    for x in range(int(x1_scaled), int(x2_scaled), 8):
        draw.line([(x, 0), (x, img.height)], fill=(0, 255, 0, 50), width=1)
    
    # Info K
    if k:
        try:
            font_info = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        except:
            font_info = ImageFont.load_default()
        draw.rectangle([5, 5, 150, 30], fill=(255, 255, 255), outline=(0, 0, 0))
        draw.text((10, 10), f"K = {k:.1f} px", fill=(0, 0, 0), font=font_info)
    
    # Zone rouge
    if highlight_code and highlight_code in zones:
        Y1, Y2 = zones[highlight_code]
        y1_scaled = Y1 * scale_y
        y2_scaled = Y2 * scale_y
        
        draw.line([(0, y1_scaled), (img.width, y1_scaled)], fill=(255, 0, 0), width=4)
        draw.line([(0, y2_scaled), (img.width, y2_scaled)], fill=(255, 0, 0), width=4)
        
        draw.rectangle(
            [x1_scaled, y1_scaled, x2_scaled, y2_scaled],
            outline=(255, 0, 0),
            width=5
        )
        
        for y in range(int(y1_scaled), int(y2_scaled), 2):
            draw.line([(x1_scaled, y), (x2_scaled, y)], fill=(255, 100, 100), width=1)
        
        try:
            font_big = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 18)
        except:
            font_big = ImageFont.load_default()
        
        text = f"CODE {highlight_code}"
        draw.text((x1_scaled + 10, y1_scaled + 5), text, fill=(255, 0, 0), font=font_big)
    
    return img

def get_saved_scope(pdf_name: str, form: str):
    """Retourne le scope manuel stocké pour (pdf_name, form) ou None."""
    return st.session_state["manual_scopes"].get((pdf_name, form))

def save_scope(pdf_name: str, form: str, X1: float, X2: float):
    """Sauvegarde/écrase le scope manuel pour (pdf_name, form)."""
    st.session_state["manual_scopes"][(pdf_name, form)] = (float(X1), float(X2))

def get_effective_scope(page, pdf_name: str, form: str):
    """
    Pour l'extraction batch : on prend le scope manuel si dispo,
    sinon le scope auto calculé.
    """
    saved = get_saved_scope(pdf_name, form)
    if saved:
        return saved
    auto = find_column_scope(page, form)
    return auto


# ============================ UI STREAMLIT ============================
# Upload multiple PDFs (max 3)
uploaded_files = st.file_uploader(
    "📁 Glissez-déposez vos PDF (3 max)",
    type=["pdf"],
    accept_multiple_files=True,
    help="Upload jusqu'à 3 PDFs pour générer un Excel sur 3 années"
)

# Upload modèle Excel (optionnel)
model_file = st.file_uploader(
    "📊 Modèle Excel",
    type=["xlsx"],
    help="Si fourni, génère un Excel avec les données extraites"
)

if uploaded_files:
    # Limiter à 3 PDFs
    if len(uploaded_files) > 3:
        st.error("❌ Maximum 3 PDF autorisés")
        uploaded_files = uploaded_files[:3]
    
    st.success(f"✅ {len(uploaded_files)} PDF chargé(s)")
    
    # Charger tous les PDFs
    pdf_data = []
    for idx, uploaded_file in enumerate(uploaded_files):
        pdf_bytes = uploaded_file.read()
        pdf_name = uploaded_file.name
        
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            total_pages = len(pdf.pages)
            pdf_data.append({
                "name": pdf_name,
                "bytes": pdf_bytes,
                "total_pages": total_pages,
                "index": idx
            })

    show_debug = st.checkbox(
    "🔍 Afficher le debug visuel (prévisualisation et réglage colonne montants)",
    value=False)

    
    if show_debug:
        st.markdown("---")
        st.markdown("### 🔍 Afficher la visualisation de la liasse fiscale")

        # Sélection du PDF à visualiser
        pdf_names = [f"{p['name']} ({p['total_pages']} pages)" for p in pdf_data]
        selected_pdf_name = st.selectbox(
            "Choisir le PDF à visualiser",
            options=pdf_names,
            key="pdf_selector"
        )
        
        selected_idx = pdf_names.index(selected_pdf_name)
        selected_pdf = pdf_data[selected_idx]
        
        st.info(f"📄 Visualisation: **{selected_pdf['name']}**")
        
        # Ouvrir le PDF sélectionné
        with pdfplumber.open(io.BytesIO(selected_pdf['bytes'])) as pdf:
            total_pages = selected_pdf['total_pages']
            page_num = st.number_input(
                "📄 Page", 
                min_value=1, 
                max_value=total_pages, 
                value=1, 
                key="page_selector"
            )
            
            page = pdf.pages[page_num - 1]
            text = page.extract_text() or ""
            form = find_form_from_text(text)
            
            st.markdown(f"### Page {page_num}/{total_pages} - Form: **{form or '❌ Non détecté'}**")
            
            if form in ["2050", "2051", "2052", "2053"]:
                # 1. Récupérer scope auto ET éventuellement scope manuel stocké
                auto_scope = find_column_scope(page, form)
                saved_scope = get_saved_scope(selected_pdf["name"], form)

                if saved_scope:
                    base_X1, base_X2 = saved_scope
                    src_label = "manuel (déjà mémorisé)"
                elif auto_scope:
                    base_X1, base_X2 = auto_scope
                    src_label = "auto détecté"
                else:
                    base_X1, base_X2 = (0.0, float(page.width))
                    src_label = "fallback plein écran"

                st.success(
                    f"🔎 Scope {src_label} : "
                    f"X1={base_X1:.1f}, X2={base_X2:.1f}, "
                    f"largeur={base_X2-base_X1:.1f}"
                )

                st.caption(
                    "Ajuste si la colonne verte ne correspond pas à la bonne colonne montants. "
                    "Ce réglage est mémorisé pour ce PDF + ce formulaire et sera utilisé pour l'extraction finale."
                )

                # 2. Sliders pour ajuster X1 / X2
                slider_col1, slider_col2 = st.columns(2)

                with slider_col1:
                    adj_X1 = st.slider(
                        "Bord gauche (X1)",
                        min_value=0.0,
                        max_value=float(page.width),
                        value=float(base_X1),
                        step=1.0,
                        key=f"adj_X1_{selected_idx}_{page_num}_{form}"
                    )

                with slider_col2:
                    adj_X2 = st.slider(
                        "Bord droit (X2)",
                        min_value=0.0,
                        max_value=float(page.width),
                        value=float(base_X2),
                        step=1.0,
                        key=f"adj_X2_{selected_idx}_{page_num}_{form}"
                    )

                # sécurité : si inversé
                if adj_X2 < adj_X1:
                    adj_X1, adj_X2 = adj_X2, adj_X1

                # Sauvegarde en session pour ce PDF + ce form
                save_scope(selected_pdf["name"], form, adj_X1, adj_X2)

                final_scope = (adj_X1, adj_X2)


                # 3. Extraction montants avec ce scope
                code_positions = find_code_positions(page, form)
                
                k = calculate_k_constant(code_positions, form)
                
                if k is None:
                    st.error("❌ Impossible de calculer K")
                else:
                    zones = calculate_zones_with_k(code_positions, k, page)
                    amounts = extract_all_amounts(page, code_positions, zones, final_scope)
                    
                    col1, col2 = st.columns([1, 2])
                    
                    with col1:
                        st.markdown("### 🔍 Code")
                        
                        codes_found = sorted(code_positions.keys())
                        highlight_code = st.selectbox(
                            "Sélectionner", 
                            options=["Aucun"] + codes_found, 
                            index=0, 
                            key=f"code_select_{page_num}"
                        )
                        
                        if highlight_code != "Aucun":
                            pos = code_positions[highlight_code]
                            Y1, Y2 = zones[highlight_code]
                            amount = amounts.get(highlight_code)
                            
                            st.markdown(f"#### 🔴 **{highlight_code}**")
                            st.markdown("**💰 Montant:**")
                            if amount:
                                if amount.startswith("-"):
                                    st.error(f"**{amount}** (négatif)")
                                else:
                                    st.success(f"**{amount}**")
                            else:
                                st.warning("Vide")
                    
                    with col2:
                        highlight = None if highlight_code == "Aucun" else highlight_code
                        img = draw_scope_visualization(
                            page,
                            final_scope,
                            code_positions,
                            zones,
                            k,
                            highlight
                        )
                        st.image(img, use_container_width=True)
                    
                    with st.expander("📊 Tableau complet"):
                        import pandas as pd
                        
                        df_data = []
                        for code_ in codes_found:
                            pos = code_positions[code_]
                            Y1, Y2 = zones[code_]
                            amount = amounts.get(code_, None)
                            df_data.append({
                                "Code": code_,
                                "Y": f"{pos['y']:.1f}",
                                "Montant": amount if amount else ""
                            })
                        
                        df = pd.DataFrame(df_data)
                        st.dataframe(df, use_container_width=True)
                        
                        amounts_found = sum(1 for a in amounts.values() if a)
                        negatives = sum(1 for a in amounts.values() if a and a.startswith("-"))
                        
                        col_m1, col_m2, col_m3 = st.columns(3)
                        col_m1.metric("Montants", f"{amounts_found} / {len(codes_found)}")
                        col_m2.metric("Négatifs", negatives)
                        col_m3.metric("Positifs", amounts_found - negatives)
            
            else:
                st.warning("⚠️ Formulaire non reconnu ou non géré")



    
    # ==================== EXTRACTION COMPLÈTE + EXCEL ====================
    st.markdown("---")
    st.markdown("### 📊 Extraction complète vers Excel")
    
    if st.button("Generez 🚀", type="primary"):
        with st.spinner("Extraction en cours..."):
            all_extractions = {}  # {year: {code: value}}
            
            for pdf_info in pdf_data:
                st.info(f"📄 Traitement: **{pdf_info['name']}**")
                
                # Extraire l'année depuis le nom du fichier
                filename = pdf_info['name']
                year_match = re.search(r'(\d{4})', filename)
                if year_match:
                    detected_year = int(year_match.group(1))
                else:
                    detected_year = None
                with pdfplumber.open(io.BytesIO(pdf_info['bytes'])) as pdf:
                    pdf_extraction = {}
                    
                    for page_idx in range(len(pdf.pages)):
                        page = pdf.pages[page_idx]
                        text = page.extract_text() or ""
                        form = find_form_from_text(text)
                        
                        if not form:
                            continue
                        
                        scope_for_batch = get_effective_scope(page, pdf_info["name"], form)
                        if not scope_for_batch:
                            continue
                        
                        code_positions = find_code_positions(page, form)
                        k = calculate_k_constant(code_positions, form)
                        
                        if not k:
                            continue
                        
                        zones = calculate_zones_with_k(code_positions, k, page)
                        amounts = extract_all_amounts(page, code_positions, zones, scope_for_batch)
                        
                        pdf_extraction.update(amounts)
                    
                    if not detected_year:
                        st.warning(f"⚠️ Année non détectée pour {pdf_info['name']}")
                        detected_year = st.number_input(
                            f"Année pour {pdf_info['name']}", 
                            min_value=2000, 
                            max_value=2100, 
                            value=2024,
                            key=f"year_{pdf_info['index']}"
                        )
                    
                    all_extractions[detected_year] = pdf_extraction
            
            for year in sorted(all_extractions.keys()):
                st.info(f"**{year}** : {len(all_extractions[year])} codes")
            
            if model_file and all_extractions:
                st.markdown("---")
                st.markdown("### 📊 Génération Excel")
                
                with st.spinner("Génération..."):
                    model_bytes = model_file.read()
                    
                    # ✅ CORRECTION DU ROUTAGE ICI
                    codes_bilan = CODES_BILAN_ALL
                    codes_cr = CODES_CR_ALL
                    
                    try:
                        excel_bytes = fill_excel(
                            model_bytes=model_bytes,
                            year_to_mapping=all_extractions,
                            bilan_sheet="Saisie Bilan",
                            cr_sheet="Saisie Cpte Res.",
                            codes_bilan=codes_bilan,
                            codes_cr=codes_cr
                        )
                        
                        st.success("✅ Excel généré!")
                        
                        years_str = "_".join(str(y) for y in sorted(all_extractions.keys()))
                        filename = f"liasse_fiscale_{years_str}.xlsx"
                        
                        st.download_button(
                            label="💾 Télécharger Excel",
                            data=excel_bytes,
                            file_name=filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    except Exception as e:
                        st.error(f"❌ Erreur: {str(e)}")
            
            elif not model_file:
                st.info("💡 Upload un modèle Excel pour générer le fichier")

else:
    st.info("Upload et Générez")
    
    with st.expander("📖 Mode d'emploi"):
        st.markdown("""
        ### 🔧 Workflow
        1. Upload 3 PDF (1 par année)
        2. Upload template Excel
        3. Possibilité d'afficher la visualition (réglage manuel)
        4. Clic "Générez"
        5. Télécharger Excel généré
        """)

